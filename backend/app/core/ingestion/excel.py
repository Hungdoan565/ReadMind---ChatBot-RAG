"""Excel (.xlsx) ingestion helpers."""

from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
import logging
from typing import Iterable, List

from langchain_core.documents import Document
from openpyxl import load_workbook

from app.config import settings

logger = logging.getLogger(__name__)


def parse_excel_bytes(
    content: bytes,
    filename: str,
    doc_id: str,
    room_code: str,
) -> List[Document]:
    """Parse an Excel workbook into LangChain documents grouped by sheet and row range."""
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    documents: List[Document] = []

    for sheet in workbook.worksheets:
        rows = sheet.iter_rows(values_only=True)
        try:
            header_row = next(rows)
        except StopIteration:
            continue

        headers = [_normalize_header(cell, idx) for idx, cell in enumerate(header_row)]
        buffered_rows: List[str] = []
        chunk_start_row: int | None = None
        current_size = 0

        for row_index, row in enumerate(rows, start=2):
            if _is_empty_row(row):
                continue

            row_text = _format_row(headers, row)
            if not row_text:
                continue

            if chunk_start_row is None:
                chunk_start_row = row_index

            projected_size = current_size + len(row_text) + 1
            if buffered_rows and projected_size > settings.CHUNK_SIZE:
                documents.append(
                    _build_sheet_document(
                        buffered_rows,
                        filename=filename,
                        doc_id=doc_id,
                        room_code=room_code,
                        sheet_name=sheet.title,
                        row_start=chunk_start_row,
                        row_end=row_index - 1,
                    )
                )
                buffered_rows = []
                current_size = 0
                chunk_start_row = row_index

            buffered_rows.append(row_text)
            current_size += len(row_text) + 1

        if buffered_rows and chunk_start_row is not None:
            row_end = (chunk_start_row - 1) + len(buffered_rows)
            documents.append(
                _build_sheet_document(
                    buffered_rows,
                    filename=filename,
                    doc_id=doc_id,
                    room_code=room_code,
                    sheet_name=sheet.title,
                    row_start=chunk_start_row,
                    row_end=row_end,
                )
            )

    logger.info(
        "Parsed Excel workbook %s into %d document block(s)", filename, len(documents)
    )
    return documents


def _normalize_header(value: object, index: int) -> str:
    if value is None:
        return f"column_{index + 1}"
    text = str(value).strip()
    return text or f"column_{index + 1}"


def _is_empty_row(row: Iterable[object]) -> bool:
    return all(cell is None or str(cell).strip() == "" for cell in row)


def _format_row(headers: List[str], row: Iterable[object]) -> str:
    cells = []
    for header, value in zip(headers, row):
        if value is None or str(value).strip() == "":
            continue
        cells.append(f"{header}: {_stringify_cell(value)}")
    return " | ".join(cells)


def _stringify_cell(value: object) -> str:
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def _build_sheet_document(
    rows: List[str],
    *,
    filename: str,
    doc_id: str,
    room_code: str,
    sheet_name: str,
    row_start: int,
    row_end: int,
) -> Document:
    return Document(
        page_content="\n".join(rows),
        metadata={
            "source": filename,
            "doc_id": doc_id,
            "room_code": room_code,
            "sheet_name": sheet_name,
            "row_range": f"{row_start}-{row_end}",
            "file_type": "xlsx",
        },
    )
